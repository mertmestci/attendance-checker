import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl import Workbook


class Student:
    def __init__(self, name ,id, department,section):
        self.name = name
        self.id = id
        self.department = department
        self.section=section

class StudentList:
    def __init__(self):
        self.students = []

    def add_student(self, student):
        self.students.append(student)

    def remove_student(self, student_info):
        for student in self.students:
            if f"{student.name}, {student.id}" == student_info:
                self.students.remove(student)
                break

    def get_students_by_section(self, section):
        return [student for student in self.students if student.section.startswith(section)]

    def load_students_from_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb.active
        first_row_skipped = False 
        for row in ws.iter_rows(values_only=True):
             if not first_row_skipped:  
                 first_row_skipped = True  
                 continue  
             if len(row) >= 4:
                
                student_name=row[1]
                name_parts = student_name.split()
                formatted_name= name_parts[-1] +' '+ name_parts[0]
                student = Student(formatted_name, row[0], row[2],row[3])
                self.add_student(student)

class GUI:
    def __init__(self, root):
        self.root = root
        self.student_list = StudentList()

        self.title_label = tk.Label(self.root, text="Attendance Keeper v1.0", font="Calibri 20 bold")

        self.title_label.grid(row=0,column=1, columnspan=6, pady=20,padx=17,sticky=tk.EW)

        self.label1 = tk.Label(self.root, text="Select student list Excel file:", font="Calibri 13 bold", anchor="w")

        self.label1.grid(row=1, column=0, columnspan=2, sticky="w", pady=0)
    
        self.section_combobox = ttk.Combobox(self.root,width=10)
        self.section_combobox.grid(row=3, column=2, sticky="ew")
        self.section_combobox.bind("<<ComboboxSelected>>", self.section_selected)

      
        self.students_listbox = tk.Listbox(self.root,height=3, selectmode=tk.MULTIPLE)
        self.students_listbox.grid(row=3, column=0, rowspan=3, columnspan=2, sticky='swen')
      
        self.attended_students_listbox = tk.Listbox(self.root,height=3,selectmode=tk.MULTIPLE)
        self.attended_students_listbox.grid(row=3, column=3, rowspan=3, columnspan=2, sticky="swen")

      
        self.add_button = tk.Button(self.root, text="add", command=self.add_students)
        self.add_button.grid(row=4, column=2, sticky="ew")

        
        self.remove_button = tk.Button(self.root, text="Remove", command=self.remove_students)
        self.remove_button.grid(row=5, column=2, sticky="ew")

      
        self.import_button = tk.Button(self.root, text="Import List", command=self.import_student_list)
        self.import_button.grid(row=1, column=2, sticky=tk.EW)

        self.label2 = tk.Label(self.root, text="Select a Student:", font="Calibri 13 bold")
        self.label2.grid(row=2, column=0, columnspan=2)

        self.label3 = tk.Label(self.root, text="Please enter week:", font="Calibri 11")
        self.label3.grid(row=6, column=2, sticky="ew")

        self.label4 = tk.Label(self.root, text="Please select file type:", font="Calibri 11 bold")
        self.label4.grid(row=6, column=0, sticky="w")

        self.entry = tk.Entry(self.root, width=18)
        self.entry.grid(row=6, column=3, sticky="ew")

        self.label5 = tk.Label(self.root, text="Section:", font="Calibri 13 bold")
        self.label5.grid(row=2, column=2)

        self.label6 = tk.Label(self.root, text="Attended Students:", font="Calibri 13 bold")
        self.label6.grid(row=2, column=3, columnspan=2, padx=25)

        self.export_button = tk.Button(self.root, text="Export file", command=self.export_attendance)
        self.export_button.grid(row=6, column=4)

        self.file_type_combobox = ttk.Combobox(self.root, values=["xlsx", "csv", "txt"])
        self.file_type_combobox.current(0)
        self.file_type_combobox.grid(row=6, column=1)


    def section_selected(self, event):
        selected_section = self.section_combobox.get()
        students = self.student_list.get_students_by_section(selected_section)
        self.students_listbox.delete(0, tk.END)
        self.attended_students_listbox.delete(0, tk.END)
        
        for student in students:
            self.students_listbox.insert(tk.END, f"{student.name}, {student.id}, {student.department}")

    def add_students(self):
        selected_indices = self.students_listbox.curselection()
        selected_students = [self.students_listbox.get(index) for index in selected_indices]
        for student_info in selected_students:
            if student_info not in self.attended_students_listbox.get(0, tk.END):
                self.attended_students_listbox.insert(tk.END, student_info)
          
        

    def remove_students(self):
        selected_indices = self.attended_students_listbox.curselection()
        selected_students = [self.attended_students_listbox.get(index) for index in selected_indices]
        for student_info in selected_students:
            self.student_list.remove_student(student_info)
            self.attended_students_listbox.delete(selected_indices[0])

    def import_student_list(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            wb = load_workbook(filename)
            ws = wb.active
            self.section_combobox['values'] = []
            unique_sections = set(cell.value for cell in ws['D'][1:] if cell.value)
            self.section_combobox['values'] = sorted(list(unique_sections))
            
            if unique_sections:
                self.section_combobox.current(0)
            self.student_list.load_students_from_excel(filename)
            self.section_selected(self)
          

    def update_students_listbox(self):
        self.students_listbox.delete(0, tk.END)
        for student in self.student_list.students:
            self.students_listbox.insert(tk.END, f"{student.surname}, {student.name}, {student.id}")

    def export_attendance(self):
        file_type = self.file_type_combobox.get()
        section_name = self.section_combobox.get()
        week = self.entry.get()
        file_name = f"{section_name} week {week}.{file_type}"
        selected_students=self.attended_students_listbox.get(0, tk.END) 
        
        if file_type.lower() == "csv":
            raise BaseException("File type is not supported")
        elif file_type.lower() == "xlsx":


            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Name", "Department"])
            for student in selected_students:
               student_data = student.split(", ")
               student = Student(student_data[0],  student_data[1], student_data[2],self.section_combobox.get())
               ws.append([student.id,student.name,student.department])
            wb.save(file_name)
        elif file_type.lower() == "txt":
            with open(file_name, "w") as file:
                file.write("ID\tName\t\tDepartment\n")
                for student in selected_students:
                    student_data = student.split(", ")
                    student = Student(student_data[0],  student_data[1], student_data[2],self.section_combobox.get())
                    file.write(f"{student.id}\t{student.name} \t{student.department}\n")

def main():
    root = tk.Tk()
    root.title('student attendance')
    root.geometry("650x250+500+350")
    root.resizable(False, False)  
    app = GUI(root)
    
    root.mainloop()
main()




